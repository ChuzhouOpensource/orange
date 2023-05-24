# import  paramiko
# from paramiko.ssh_exception import AuthenticationException, SSHException,NoValidConnectionsError
# if __name__ == '__main__':
#     ssh = paramiko.SSHClient()
#
#     ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy)
#
#     try:
#         ssh.connect(hostname='10.211.55.8', port=22, username='makun1', password='1', timeout=10)
#
#         stdin, stdout, stderr = ssh.exec_command('hostname')
#
#         result = stdout.read()
#         print(result.decode('utf-8'))
#
#     except NoValidConnectionsError as e:
#         print(f'请输入正确的用户名或密码1:{e}')
#     except SSHException as e:
#         print(f'请输入正确的用户名或密码2:{e}')
#
#     except AuthenticationException as e:
#         print(f'请输入正确的用户名或密码3:{e}')
#
#     except Exception as e:
#         print(f'登录地址或端口有误:{e}')
#
# a = 1
# print(1) if isinstance(a,str) else a
from openpyxl import load_workbook
# from   orange_api.apps.host.models import Host
#
# host_list = Host.objects.values_list('id', 'name')
# print(host_list)
# 加载某一个excel文件
wb = load_workbook('./主机导入模版.xlsx')
# 获取worksheet对象的两种方式
worksheet = wb.worksheets[0]


for row in worksheet.iter_rows(2):
    print(row[5].value, type(row[5].value))

