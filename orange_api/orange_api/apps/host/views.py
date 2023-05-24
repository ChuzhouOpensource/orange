from rest_framework.generics import ListCreateAPIView
from rest_framework.viewsets import ModelViewSet,ReadOnlyModelViewSet
from  rest_framework.views import APIView
from rest_framework.response import Response
from .models import  Host, HostCategory
from .serializers import HostModelSerializer, HostCategoryModelSerializer

from openpyxl import load_workbook
from io import  BytesIO


class HostCategoryApiView(ListCreateAPIView):
    """主机类别视图类"""

    queryset = HostCategory.objects.filter(is_show=True,is_deleted=False).order_by("orders", "-id")
    serializer_class =  HostCategoryModelSerializer

class HostApiView(ModelViewSet):
    '''主机信息'''
    queryset = Host.objects.all()
    serializer_class = HostModelSerializer

def read_host_excel_data(io_data,default_password):
    """
        从excel中读取主机列表信息
        io_data: 主机列表的字节流
        default_password: 主机的默认登录密码
    """
    # 加载某一个excel文件
    wb = load_workbook(io_data)

    # 获取worksheet对象的两种方式
    worksheet = wb.worksheets[0]


    # 查询出数据库现有的所有分类数据[ID，name]
    # 由于拿到的是分类名称，所以我们要找到对应名称的分类id，才能去数据库里面存储
    category_list = HostCategory.objects.values_list('id', 'name')


    #接收主机的登录信息
    host_info_list = []
    for row in worksheet.iter_rows(2):
        if not row[0].value : continue
        host_info_dict = {}

        for category in category_list:
            if  row[0].value.strip() == category[1]:
                host_info_dict['category'] = category[0]
                break

        host_info_dict['name'] = row[1].value #主机名称
        host_info_dict['ip_addr'] = row[2].value #主机地址
        host_info_dict['port'] = row[3].value #登录端口
        host_info_dict['username'] = row[4].value #登录用户
        host_info_dict['description'] = row[6].value #备注

        #为空，接收defalut_password
        password = row[5].value
        if not password:
            host_info_dict['password'] = default_password

        host_info_dict['password'] = password

        #将获得主机信息已dict形式加到host_info_list列表中
        host_info_list.append(host_info_dict)


    #校验主机数据
    respone_data = {}
    serializers_host_res_data = [] #已通过验证的数据礼拜
    res_error_data = [] #验证错误的内容
    for k, host_data in enumerate(host_info_list):

        host_serializers = HostModelSerializer(data=host_data)
        if host_serializers.is_valid(raise_exception=True):
            #通过验证的数据，反序列化
            vali_data = host_serializers.save()
            serializers_host_res_data.append(vali_data) #将对应的对象添加至列表

        else:
            res_error_data.append(f'execl表格第{k}行数据错误，请修改后再次提交！')


    #序列化返回数据
    serializers = HostModelSerializer(instance=serializers_host_res_data, many=True)

    respone_data['data'] = serializers.data
    respone_data['error'] = res_error_data

    return respone_data


class ExeclHostView(APIView):
    def post(self, request):
        host_excel = request.FILES.get('host_excel')
        default_password = request.data.get('default_password')

        #  把上传文件全部写入到字节流，就不需要保存到服务端硬盘了。
        io_data = BytesIO()
        for line in host_excel:
            io_data.write(line)
        #
        data = read_host_excel_data(io_data, default_password)
        #
        return Response(data)



